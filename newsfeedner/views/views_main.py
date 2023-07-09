import json
import logging
import os
import sys
from datetime import datetime
from time import time
from typing import List, Tuple
import uuid

import pytz
import requests
from django.contrib.staticfiles.storage import staticfiles_storage
from django.db import transaction
from django.db.models import Q, QuerySet
from django.db.utils import IntegrityError
from django.shortcuts import get_object_or_404
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import permissions
from rest_framework import status
from rest_framework.generics import ListAPIView
from rest_framework.response import Response

# Добавлено 03.02.2023
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import serializers

from newsfeedner.models import (
    TradeEvent,
    TradeEventRelevant,
    TradeEventForApproval,
)
from newsfeedner.permissions import ApproverPermissions
from newsfeedner.serializers.serializers_main import (
    EventSerializerRead,
    EventApprovalSerializerWrite,
    EventApprovalSerializerRead,
    EventRelevantSerializerRead,
    EventRelevantSerializerWrite,
    EmbeddingSerializer,
    EventSerializerWrite,
    EventApprovalSerializerWriteStatus,
)
from newsfeedner.serializers.serializers_post_ids import (
    CheckedIDSerializer,
    ApprovedIDSerializer,
)

# from newsfeedner.utils.queryset_process import (
#     process_queryset_classes,
#     process_text_locations,
#     get_article_abstract,
# )
from newsfeedner.utils.trade_utils import query_regions_dict as query_regions
from newsfeedner.utils.trade_utils import smtk_products

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

stdout_handler = logging.StreamHandler(sys.stdout)
stdout_handler.setLevel(logging.DEBUG)
stdout_handler.setFormatter(formatter)

file_handler = logging.FileHandler("logs.log")
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(stdout_handler)

default_annotation_path = staticfiles_storage.path("newsfeedner/annotation_config.json")
with open(default_annotation_path) as f:
    default_config = json.load(f)
annotation_config = None
if not annotation_config:
    annotation_config = default_config

EMBEDDING_SERVICE = os.getenv("EMBEDDING_SERVICE")


class EventList(ListView):
    context_object_name = "mainpage_data"
    template_name = "newsfeedner/events.html"
    paginate_by = 10
    n_entities = 5

    # def filter_context_data(self, queryset): #, region="",country=""
    # to_del_ids = set()
    # for q in queryset:
    # location_str = q.locations
    # locations, regions = process_text_locations(location_str, region=region, country=country)

    # q.classes = "; ".join(q.classes)
    # q.classes = "; ".join(sorted(set(q.classes.split("; "))))

    #     title = q.title
    #     title, sents = get_article_abstract(title, locations)
    #     locations_filter = [sc.lower()[:-1] for sc in locations]
    #     if not sents and not any(sc in title.lower()
    #                              for sc in locations_filter):
    #         to_del_ids.add(q.id)
    #         continue
    #     q.title = title

    #     location_str = ", ".join(locations)
    #     q.locations = location_str

    # queryset = [q for q in queryset]  #if q.id not in to_del_ids
    # return queryset

    def filter_queryset_by_location(self, queryset, region=None, country=None):
        """
        Функция получает на вход queryset и фильтры по региону и стране.
        Фильтрует queryset по заданным фильтрам и возвращает новый queryset
        region - value. Например, "Страны_Африки"
        country - value. Например, "Коморские_острова",
        """
        if country:
            country = country.replace("_", " ")
            country = country.replace("Все страны", "")
        if not country and region.startswith("Страны"):
            region_list = query_regions.get(region)
            if "Африки" in region:
                region_strings = ["африк", "брикс", "опек"]
            elif "Латинской" in region:
                region_strings = ["латинск", "америк", "опек"]
            elif "СНГ" in region:
                region_strings = ["снг", "брикс", "еаэс"]
            elif "Европ" in region or "США" in region:
                region_strings = ["европ", "сша", "америк", "ec"]
            elif "Океани" in region or "Австрал" in region:
                region_strings = ["океан", "австрал"]
            elif "Азии" in region:
                region_strings = ["азии", "азия", "асеан", "брикс", "опек"]
            elif "Ближн" in region:
                region_strings = ["восток", "ближний", "опек"]
            else:
                region_strings = ""

            combined_query = Q(locations__icontains=region_strings[0])
            for region_string in region_strings[1:]:
                combined_query &= Q(locations__icontains=region_string)
            for country in region_list:
                if len(country) > 4:
                    country = country[:-1]
                # швейцария contains цар
                if country == "ЦАР":
                    combined_query |= Q(locations__contains=country.upper())
                # боливия contains ливия
                elif country == "ливи":
                    combined_query |= Q(locations__icontains=" " + country)
                    combined_query &= ~Q(locations__icontains="боливия")
                else:
                    combined_query |= Q(locations__icontains=country)
            queryset = queryset.filter(combined_query)

        elif country:
            # "match abbreviated countries only in the original cases"
            # "ЦАР not -> Швейцария"
            if country.isupper():
                queryset = queryset.filter(locations__contains=country)
            else:
                queryset = queryset.filter(locations__icontains=country.lower())
        return queryset

    def filter_queryset_by_relation(self, queryset, relation=None):
        """
        Функция получает на вход queryset и фильтры по типу отношений.
        Фильтрует queryset по заданным фильтрам и возвращает новый queryset
        relation - value. Например, "Внешняя_торговля"
        """
        if relation:
            relation = relation.replace("Все типы", "")
        if relation:
            query_relations = [relation]
            relation_filter = Q(classes__contains=query_relations[:1])

            for query_relation in query_relations[1:]:
                relation_filter |= Q(classes__contains=[query_relation])
            queryset = queryset.filter(relation_filter)

        return queryset

    def filter_queryset_by_product(self, queryset, product_branch=None, product=None):
        """
        Функция получает на вход queryset и фильтры по СМТК-коду товара и/ или товарного раздела.
        Фильтрует queryset по заданным фильтрам и возвращает новый queryset
        product_branch - value. Например, "0_-_Пищевые_продукты_и_живые_животные"
        product - value. Например, "01_-_Мясо_и_мясопродукты"
        """
        if product_branch or product:
            product_branch = product_branch.replace("Все товарные разделы", "")
            product_branch = product_branch.replace("_", " ")
            product = product.replace("_", " ")
            product = product.replace("Все товарные группы", "")

        if product:
            query = Q(itc_codes__contains=product)
            queryset = queryset.filter(query)

        elif not product and product_branch:
            products = smtk_products[product_branch]

            query = Q(itc_codes__contains=product_branch)
            for product in products:
                query |= Q(itc_codes__contains=product)
            queryset = queryset.filter(query)
        return queryset

    def get_queryset(self):
        print("start")
        queryset = TradeEvent.objects.all()
        if self.request.method == "GET":
            region = self.request.GET.get("region", "")
            country = self.request.GET.get("country", "")
            relation = self.request.GET.get("relation", "")
            product_branch = self.request.GET.get("product_branch", "")
            product = self.request.GET.get("product", "")

            queryset = self.filter_queryset_by_location(queryset, region, country)
            queryset = self.filter_queryset_by_relation(queryset, relation)
            queryset = self.filter_queryset_by_product(queryset, product_branch, product)
        return queryset


# Base class
class EventApiView(ListAPIView, EventList):
    """
    API возвращает список событий из таблицы trade_news_events, в которой хранятся данные,
    обработанные моделями.
    И записывает данные в таблице trade_news_for_approval
    """

    read_serializer_class = EventSerializerRead
    write_event_serializer_class = EventSerializerWrite
    write_approval_serializer_class = EventApprovalSerializerWrite
    checkedids_serializer = CheckedIDSerializer
    embedding_serializer = EmbeddingSerializer

    permission_classes = (permissions.IsAuthenticated,)
    region = openapi.Parameter(
        "region",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="название региона, например: 'Страны Африки'. " "Cписок всех стран и регионов можно получить " "с помощью метода 'get_regions'.",
    )

    relation = openapi.Parameter(
        "relation",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="тип отношения в новости, например: "
        "'Внешняя торговля'"
        "список всех товарных групп можно получить "
        "с помощью метода 'get_relations'. "
        "Поиск реализован как поиск по тексту, поэтому в теории "
        "можно передавать и просто 'торговля'",
    )

    country = openapi.Parameter(
        "country",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="название страны, например: 'Египет'. " "Cписок всех стран можно получить " "с помощью метода 'get_countries'.",
    )

    product_branch = openapi.Parameter(
        "product_branch",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="название товарного раздела СМТК, например: "
        "'0 - Пищевые продукты и живые животные'. "
        "Cписок всех товарных разделов можно получить "
        "с помощью метода 'get_products'. "
        "Поиск реализован как поиск по тексту, поэтому в теории "
        "можно передавать и просто 'напитки' (но не рекомендуется)",
    )

    product = openapi.Parameter(
        "product",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="название товарной группы СМТК, например: "
        "'11 - Напитки'. "
        "Cписок всех товарных групп можно получить "
        "с помощью метода 'get_products'. "
        "Поиск реализован как поиск по тексту, поэтому в теории "
        "можно передавать и просто 'напитки' (но не рекомендуется)",
    )

    start_date = openapi.Parameter(
        "start_date",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="дата начала периода, например: 2023-01-01",
    )

    end_date = openapi.Parameter(
        "end_date",
        in_=openapi.IN_QUERY,
        type=openapi.TYPE_STRING,
        description="дата окончания периода, например: 2023-03-08",
    )

    def check_dates(self, dates):
        dates = dates.split(",")
        for d in dates:
            try:
                d = datetime.strptime(d, "%Y-%m-%d")
            except ValueError:
                raise serializers.ValidationError({"dates": f"'{d}' date format is wrong. Date should be in format YYYY-MM-DD"})

    def get_query_params(self) -> Tuple[Tuple, List]:
        """
        Функция получает параметры запроса и возвращает список имен и список значений параметров
        """
        fields = ("region", "relation", "country", "product_branch", "product", "start_date", "end_date")
        values = []
        for field in fields:
            value = self.request.query_params.get(field)
            if not value:
                value = self.kwargs.get(field)
            if not value:
                value = ""
            values.append(value)
        return fields, values

    def get_queryset(self, objectTable) -> QuerySet:
        """
        Получает из таблицы trade_news_events сырые данные фильтрует их по параметрам
         и возвращает отфильтрованый queryset.
        Метод post позволяет записать данные в базу в таблицу trade_news_for_approval
         (модель TradeEventForApproval)
        """
        # queryset = TradeEvent.objects.all()
        queryset = objectTable.objects.filter(dates__range=[self.start_date, self.end_date]).order_by("-dates")

        fields, values = self.get_query_params()

        queryset = self.filter_queryset_by_location(queryset, values[0], values[2])
        queryset = self.filter_queryset_by_relation(queryset, values[1])
        queryset = self.filter_queryset_by_product(queryset, values[3], values[4])
        return queryset

    # def list(self, request, *args, **kwargs):
    #     start_time = time()
    #     queryset = self.get_queryset(TradeEvent)
    #     # fields, values = self.get_query_params()
    #     # region = values[0]
    #     # relation = values[1]
    #     # country = values[2]
    #     # product = values[4]

    #     # queryset = self.filter_context_data(queryset) #, region, country

    #     page = self.paginate_queryset(queryset)
    #     if page is not None:
    #         queryset = page
    #     serializer = self.get_serializer(queryset, many=True)
    #     events = serializer.data
    #     data = dict()
    #     # ugly but works
    #     data["count"] = self.paginator.page.paginator.num_pages
    #     data["next"] = self.paginator.get_next_link()
    #     data["previous"] = self.paginator.get_previous_link()
    #     data["results"] = events
    #     print("--- %s seconds ---" % (time() - start_time))
    #     return Response(data)

    @csrf_exempt
    @method_decorator(
        swagger_auto_schema(
            manual_parameters=[
                region,
                relation,
                country,
                product_branch,
                product,
                start_date,
                end_date,
            ],
        ),
        name="list",
    )
    def get(self, request, *args, **kwargs):
        self.start_date = self.request.GET.get("start_date", "")
        self.end_date = self.request.GET.get("end_date", "")
        try:
            self.check_dates(self.start_date)
            self.check_dates(self.end_date)
        except ValueError as e:
            print(e)
            return f"Error {e}"
        return super().get(request, *args, **kwargs)


# Страница для всех пользователей - релевантные новости
class EventApiViewRelevant(EventApiView):
    """
    Класс предоставляет данные только на чтение из таблицы trade_news_relevant.
    """

    http_method_names = ["get", "head", "options", "trace"]
    read_serializer_class = EventRelevantSerializerRead
    permission_classes = (permissions.AllowAny,)

    region = EventApiView.region
    relation = EventApiView.relation
    country = EventApiView.country
    product_branch = EventApiView.product_branch
    product = EventApiView.product
    start_date = EventApiView.start_date
    end_date = EventApiView.end_date

    def get_serializer_class(self):
        return self.read_serializer_class

    @csrf_exempt
    @method_decorator(
        swagger_auto_schema(
            manual_parameters=[
                region,
                relation,
                country,
                product_branch,
                product,
                start_date,
                end_date,
            ],
        ),
        name="list",
    )
    def list(self, request, *args, **kwargs):
        queryset = super().get_queryset(TradeEventRelevant)
        page = self.paginate_queryset(queryset)
        if page is not None:
            queryset = page
        serializer = self.get_serializer(queryset, many=True)
        events = serializer.data

        data = dict()
        # ugly but works
        data["count"] = self.paginator.page.paginator.num_pages
        data["next"] = self.paginator.get_next_link()
        data["previous"] = self.paginator.get_previous_link()
        data["results"] = events
        return Response(data)


# Добавлено 07.02.2023 получить релевантные новости в формате xlsx
class EventApiViewRelevant_to_xlsx(EventApiView):
    """Класс возвращает файл в формате xlsx в соответсвии с установленными значениями фильтров."""

    http_method_names = ["get"]
    read_serializer_class = EventRelevantSerializerRead
    permission_classes = (permissions.AllowAny,)

    region = EventApiView.region
    relation = EventApiView.relation
    country = EventApiView.country
    product_branch = EventApiView.product_branch
    product = EventApiView.product
    start_date = EventApiView.start_date
    end_date = EventApiView.end_date

    @csrf_exempt
    @method_decorator(
        swagger_auto_schema(
            manual_parameters=[
                region,
                relation,
                country,
                product_branch,
                product,
                start_date,
                end_date,
            ],
            responses={"200": "Возвращает выборку в виде файла формата xlsx для заданных значений входных параметров."},
        ),
        name="Not return data",
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        queryset = super().get_queryset(TradeEventRelevant)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename={date}-trade.xlsx".format(
            date=datetime.now().strftime("%Y-%m-%d"),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = "trade_news"
        columns = ["Тип отношений", "Коды СМТК", "Страны", "Событие", "Источник", "Дата"]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="d2d2d7",
                end_color="d2d2d7",
                fill_type="solid",
            )
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            if col_num == 6:
                column_dimensions.width = 17
            else:
                column_dimensions.width = 45

        # Iterate through all rows
        for trade_news in queryset:
            row_num += 1
            # Define the data for each cell in the row
            row = [
                "; ".join(trade_news.classes),
                trade_news.itc_codes,
                trade_news.locations,
                trade_news.title,
                trade_news.url,
                trade_news.dates,
            ]
            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(response)
        return response

    # Конец добавлено


# Изменено 07.05.2023
# Страница для аннотаторов
class NewsApiView(EventApiView):
    """
    API возвращает список событий из таблицы trade_news_events, в которой хранятся данные,
    обработанные моделями в диапазоне дат начиная со "start_date" по "end_date".
    И записывает данные в таблице trade_news_for_approval
    """

    region = EventApiView.region
    relation = EventApiView.relation
    country = EventApiView.country
    product_branch = EventApiView.product_branch
    product = EventApiView.product
    start_date = EventApiView.start_date
    end_date = EventApiView.end_date

    request = ""
    read_serializer_class = EventSerializerRead
    write_event_serializer_class = EventSerializerWrite
    write_approval_serializer_class = EventApprovalSerializerWrite
    checkedids_serializer = CheckedIDSerializer
    embedding_serializer = EmbeddingSerializer

    permission_classes = (permissions.IsAuthenticated,)

    def get_serializer_class(self):
        if self.request.method == "POST":
            return self.write_approval_serializer_class
        elif self.request.method == "GET":
            return self.read_serializer_class
        else:
            return self.serializer_class

    def get_queryset(self):
        query = Q(status="not_seen")
        queryset = super().get_queryset(TradeEvent)
        queryset &= queryset.filter(query)
        return queryset

    @csrf_exempt
    @method_decorator(
        swagger_auto_schema(
            manual_parameters=[
                region,
                relation,
                country,
                product_branch,
                product,
                start_date,
                end_date,
            ],
        ),
        name="list",
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            queryset = page
        serializer = self.get_serializer(queryset, many=True)
        events = serializer.data

        data = dict()
        # ugly but works
        data["count"] = self.paginator.page.paginator.num_pages
        data["next"] = self.paginator.get_next_link()
        data["previous"] = self.paginator.get_previous_link()
        data["results"] = events
        return Response(data)

    def post_article_ids(self, request):
        """
        Метод сохраняет значения article_ids в checked_ids при сохранении данных в таблицу trade_news_for_approval
        Return: serializer.data
        """

        ids = request.data.get("article_ids")
        if ids:
            ids = ids.split(",")
        req_data = [{"checked_id": v} for v in ids]
        serializer = self.checkedids_serializer(data=req_data, many=True)
        serializer.is_valid(raise_exception=True)
        return serializer

    def post_text_embedding(self, request, uuid):
        """
        Метод получает вектор из текста через внешнее API.
        ID новости добавляется в методе post
        Return: данные записи в trade_news_embeddings без id

        """
        text = request.data.get("title")
        article_ids = request.data.get("article_ids")
        result = requests.post(EMBEDDING_SERVICE, json={"text": text}).json()
        embedding_data = {
            "id": uuid,
            "embedding": result["embedding"],
            "article_id": article_ids,
            "model": result["model"],
            "date_added": datetime.now(pytz.utc),
        }
        serializer = self.embedding_serializer(data=embedding_data)
        serializer.is_valid(raise_exception=True)

        return serializer

    def status_update(self, pk):
        status_update = {"status": "checked"}
        obj = TradeEvent.objects.get(pk=pk)
        serializer = self.write_event_serializer_class(obj, data=status_update, partial=True)
        if serializer.is_valid():
            return serializer
            # return a meaningful error response
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        article_ids_serializer = self.post_article_ids(request)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            try:
                obj = serializer.save()
                article_ids_serializer.save()
                emb_serializer = self.post_text_embedding(request, obj.id)
                emb_serializer.save()
                event_serializer = self.status_update(obj.id)
                event_serializer.save()

            except IntegrityError as err:
                if "newsfeedner_checkedids_pkey" in str(err):
                    raise IntegrityError("ID этой новости уже есть в базе")
                elif "unique_approval_news" in str(err):
                    raise IntegrityError("Новость с аналогичными значениями (Тип отношений, Страны, Код товара, Текст новости) уже существует")
                elif "approval_pk" in str(err):
                    raise IntegrityError({"error": "Новость с таким id уже есть в базе"})

        return Response(
            {
                "data": [
                    serializer.data,
                    article_ids_serializer.data,
                    emb_serializer.data,
                ]
            },
            status=status.HTTP_200_OK,
        )


# Страница для проверяющих
class NewsApiViewApproval(EventApiView):  # (ListAPIView, NewsList): #
    """
    Класс предоставляет данные на чтение из таблицы trade_news_for_approval
    и запись в таблицу trade_news_relevant.
    При передаче id равным пустой строке, можно добавлять новость в таблицу trade_news_relevant.
    В этом случае следующие параметры являются необязательными:
    article_ids, user_checked, date_checked, user_approved, date_approved
    """

    region = EventApiView.region
    relation = EventApiView.relation
    country = EventApiView.country
    product_branch = EventApiView.product_branch
    product = EventApiView.product
    start_date = EventApiView.start_date
    end_date = EventApiView.end_date

    http_method_names = ["get", "post", "head", "options", "trace"]
    read_serializer_class = EventApprovalSerializerRead
    write_approval_serializer_class = EventApprovalSerializerWriteStatus
    write_relevant_serializer_class = EventRelevantSerializerWrite
    approved_ids_serializer = ApprovedIDSerializer
    embedding_serializer = EmbeddingSerializer
    permission_classes = (
        permissions.AllowAny,
        ApproverPermissions,
    )

    def get_serializer_class(self):
        if self.request.method == "POST":
            return self.write_relevant_serializer_class
        elif self.request.method == "GET":
            return self.read_serializer_class
        else:
            return self.serializer_class

    def get_queryset(self):
        """
        Получает из базы сырые данные фильтрует их по полю status = checked  и возвращает отфильтрованый queryset
        """
        # queryset = TradeEventForApproval.objects.all()
        # Добавлено 16.02.2023
        # queryset = TradeEvent.objects.all()
        query = Q(status="checked")
        queryset = super().get_queryset(TradeEventForApproval)
        queryset &= queryset.filter(query)
        # Конец добавлено
        return queryset

    @csrf_exempt
    @method_decorator(
        swagger_auto_schema(
            manual_parameters=[
                region,
                relation,
                country,
                product_branch,
                product,
                start_date,
                end_date,
            ],
        ),
        name="list",
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            queryset = page
        serializer = self.get_serializer(queryset, many=True)
        events = serializer.data

        data = dict()
        # ugly but works
        data["count"] = self.paginator.page.paginator.num_pages
        data["next"] = self.paginator.get_next_link()
        data["previous"] = self.paginator.get_previous_link()
        data["results"] = events
        return Response(data)

    def post_article_ids(self, request):
        """
        Метод сохраняет значения article_ids в approved_ids при сохранении данных в таблицу trade_news_relevant
        Return: serializer.data
        """
        ids = request.data.get("article_ids")
        if ids:
            ids = ids.split(",")
        req_data = [{"approved_id": v} for v in ids]
        serializer = self.approved_ids_serializer(data=req_data, many=True)
        serializer.is_valid(raise_exception=True)
        return serializer

    def post_text_embedding(self, request, uuid):
        """
        Метод получает вектор из текста через внешнее API.
        ID новости добавляется в методе post
        Return: данные записи в trade_news_embeddings без id

        """
        text = request.data.get("title")
        article_ids = request.data.get("article_ids")
        result = requests.post(EMBEDDING_SERVICE, json={"text": text}).json()
        embedding_data = {
            "id": uuid,
            "embedding": result["embedding"],
            "article_id": article_ids,
            "model": result["model"],
            "date_added": datetime.now(pytz.utc),
        }
        serializer = self.embedding_serializer(data=embedding_data)
        serializer.is_valid(raise_exception=True)
        return serializer

    def status_update(self, pk):
        status_update = {"status": "approved"}

        obj_from_approval = TradeEventForApproval.objects.get(pk=pk)

        serializer_approval = self.write_approval_serializer_class(obj_from_approval, data=status_update, partial=True)
        if serializer_approval.is_valid():
            return serializer_approval
            # return a meaningful error response
        return Response(serializer_approval.errors, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        if not request.data["id"] == "":
            article_ids_serializer = self.post_article_ids(request)
            serializer.is_valid(raise_exception=True)
            with transaction.atomic():
                try:
                    obj = serializer.save()
                    article_ids_serializer.save()
                    approval_serializer = self.status_update(obj.id)
                    approval_serializer.save()
                except IntegrityError as err:
                    if "newsfeedner_approvedids_pkey" in str(err):
                        raise IntegrityError({"error": "ID этой новости уже есть в базе"})
                    elif "unique_relevant_news" in str(err):
                        raise IntegrityError({"error": "Новость с аналогичными значениями (Тип отношений, Страны, Код товара, Текст новости) уже существует"})
                    elif "relevant_pk" in str(err):
                        raise IntegrityError({"error": "Новость с таким id уже есть в базе"})

            return Response(
                {
                    "data": [
                        serializer.data,
                        article_ids_serializer.data,
                        # embedding_serializer.data,
                    ]
                },
                status=status.HTTP_200_OK,
            )
        else:
            # Добавлено 16.02.2023 возможность добавлять новость с сайта. Передача пустого значения Id.
            with transaction.atomic():
                try:
                    date_oper = datetime.now(pytz.utc)
                    username = request.user.username
                    relevant = TradeEventRelevant()
                    news_data = {
                        "id": uuid,
                        "classes": request.data.get("classes"),
                        "itc_codes": request.data.get("itc_codes"),
                        "locations": request.data.get("locations"),
                        "title": request.data.get("title"),
                        "url": request.data.get("url"),
                        "dates": request.data.get("dates"),
                        "article_ids": "0",
                        "product": request.data.get("product"),
                    }
                    write_relevant_serializer = self.write_relevant_serializer_class(data=news_data)
                    data_output = write_relevant_serializer.validate(news_data)
                    relevant.id = uuid.uuid4()
                    relevant.classes = data_output["classes"]
                    relevant.itc_codes = data_output["itc_codes"]
                    relevant.locations = data_output["locations"]
                    relevant.title = data_output["title"]
                    relevant.url = data_output["url"]
                    relevant.dates = data_output["dates"]
                    relevant.product = data_output["product"]
                    relevant.user_checked = username
                    relevant.date_checked = date_oper
                    relevant.user_approved = username
                    relevant.date_approved = date_oper
                    relevant.save()
                    # write_relevant_serializer.save()

                    embedding_serializer = self.post_text_embedding(request, relevant.id)
                    embedding_serializer.save()

                except IntegrityError as err:
                    if "newsfeedner_approvedids_pkey" in str(err):
                        raise IntegrityError({"error": "ID этой новости уже есть в базе"})
                    elif "unique_relevant_news" in str(err):
                        raise IntegrityError({"error": "Новость с аналогичными значениями (Тип отношений, Страны, Код товара, Текст новости) уже существует"})
                    elif "relevant_pk" in str(err):
                        raise IntegrityError({"error": "Новость с таким id уже есть в базе"})
            return Response(
                {"result": "Новость успешно добавлена"},
                status=status.HTTP_200_OK,
            )
            # Конец добавлено
